import os
import pandas as pd
import customtkinter as ctk
from tkinter import filedialog, messagebox, Toplevel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
from tkcalendar import Calendar
from tktimepicker import AnalogPicker, timepicker, AnalogThemes

# Initialiser l'application
ctk.set_appearance_mode("System")  # Options : "Light", "Dark", "System"
ctk.set_default_color_theme("blue")  # Options : "blue", "green", "dark-blue"

class Application(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Traitement Excel avec Calculs de Temps")
        self.geometry("800x600")

        # Variables
        self.fichier_excel = None
        self.production_par_heure = None
        self.date_heure_debut = None
        self.prefixe_fichier = None

        # Interface graphique
        self.label_instruction = ctk.CTkLabel(self, text="Logiciel de traitement", font=("Arial", 16))
        self.label_instruction.pack(pady=10)

        self.button_choisir_fichier = ctk.CTkButton(self, width=250, text="Choisir un fichier Excel (.xls .xlsx .xlsm et .xlsb)", command=self.demander_fichier)
        self.button_choisir_fichier.pack(pady=10)

        self.label_production = ctk.CTkLabel(self, text="Production par heure (UVC/h):", font=("Arial", 14))
        self.label_production.pack(pady=5)

        self.entry_production = ctk.CTkEntry(self, placeholder_text="Entrez une valeur", width=250)
        self.entry_production.pack(pady=5)

        # self.label_date_heure = ctk.CTkLabel(self, text="Date/Heure de début:", font=("Arial", 14))
        # self.label_date_heure.pack(pady=5)

        self.button_select_date = ctk.CTkButton(self, text="Sélectionner la date et l'heure", command=self.ouvrir_calendrier , width=250)
        self.button_select_date.pack(pady=5)

        self.entry_date_heure = ctk.CTkEntry(self, placeholder_text="YYYY-MM-DD HH:MM:SS", width=250)
        self.entry_date_heure.pack(pady=5)

        # self.entry_heure_pauses = ctk.CTkLabel(self, text="Heure de pauses (HH:MM-HH:MM,HH:MM-HH:MM,...):", font=("Arial", 14))
        # self.entry_heure_pauses.pack(pady=5)

        self.button_select_date = ctk.CTkButton(self, text="Sélectionner les heures de pauses", command=self.pausescomp , width=250)
        self.button_select_date.pack(pady=5)    

        self.entry_heure_pauses = ctk.CTkEntry(self, placeholder_text="Entrez les pauses", width=250)
        self.entry_heure_pauses.pack(pady=5)

        self.label_prefixe = ctk.CTkLabel(self, text="Préfixe pour le fichier final :", font=("Arial", 14))
        self.label_prefixe.pack(pady=5)

        self.entry_prefixe = ctk.CTkEntry(self, placeholder_text="Ex : Résultat_", width=250)
        self.entry_prefixe.pack(pady=5)

        self.button_lancer = ctk.CTkButton(self, width=250, text="Lancer le traitement", command=self.lancer_traitement)
        self.button_lancer.pack(pady=20)

        self.button_preview = ctk.CTkButton(self, width=250, text="Aperçu du fichier", command=self.preview_fichier)
        self.button_preview.pack(pady=10)

        self.label_copyright = ctk.CTkLabel(self, text="© Yanis Bordonado - Tous droits réservés", font=("Arial", 12), anchor="e")
        self.label_copyright.pack(side="bottom", anchor="se", pady=10, padx=10)

    def ouvrir_calendrier(self):
        def choisir_date():
            date_selectionnee = cal.selection_get().strftime("%Y-%m-%d")
            self.entry_date_heure.insert(0, date_selectionnee + " ")
            fenetre_calendrier.destroy()
            self.ouvrir_horloge()

        fenetre_calendrier = Toplevel(self)
        fenetre_calendrier.title("Sélectionner une date")
        cal = Calendar(fenetre_calendrier, date_pattern="yyyy-mm-dd")
        cal.pack(pady=10)
        btn_choisir = ctk.CTkButton(fenetre_calendrier, text="Valider", command=choisir_date)
        btn_choisir.pack(pady=10)

    
        

    def ouvrir_horloge(self):
        def choisir_heure():
            heure_selectionnee = horloge.time()  # Retourne un tuple (hour, minute, period)
            print("Valeur sélectionnée :", heure_selectionnee)  # Debugging
            
            heure, minute, periode = heure_selectionnee
            
            # Ajuster l'heure en fonction de la période
            if periode == 'Après-midi' and heure != 12:  # 12h PM reste inchangé, sinon ajoute 12
                heure += 12
            elif periode == 'Matin' and heure == 12:  # 12h AM devient 00h
                heure = 0
            
            # Formater l'heure en HH:MM:SS
            heure_formatee = f"{heure:02}:{minute:02}:00"
            
            # Insérer l'heure formatée dans le champ
            self.entry_date_heure.insert(len(self.entry_date_heure.get()), heure_formatee)
            
            # Fermer la fenêtre de sélection de l'heure
            fenetre_horloge.destroy()

            

        fenetre_horloge = Toplevel(self)
        fenetre_horloge.title("Sélectionner l'heure")
        horloge = AnalogPicker(fenetre_horloge)
        horloge.pack(pady=10)
        btn_choisir = ctk.CTkButton(fenetre_horloge, text="Valider", command=choisir_heure)
        btn_choisir.pack(pady=10)
        theme = AnalogThemes(horloge)
        theme.setLightBlueTheme()  # Exemple : Thème "Lightblue"

    def pausescomp(self):
    
        self.horloge_count = 0  # Compteur pour suivre combien de fenêtres ont été ouvertes

        def ouvrir_heurloge():
            if self.horloge_count >= 2:
                return  # Ne rien faire si deux horloges ont déjà été ouvertes

            def choisir_heure():
                heure_selectionnee = horloge.time()  # Retourne un tuple (hour, minute, period)
                print("Valeur sélectionnée :", heure_selectionnee)  # Debugging
                
                heure, minute, periode = heure_selectionnee
                
                # Ajuster l'heure en fonction de la période
                if periode == 'Après-midi' and heure != 12:  # 12h PM reste inchangé, sinon ajoute 12
                    heure += 12
                elif periode == 'Matin' and heure == 12:  # 12h AM devient 00h
                    heure = 0
                
                # Formater l'heure en HH:MM:SS
                heure_formatee = f"{heure:02}:{minute:02}"
                
                # Ajouter une virgule si le champ n'est pas vide, mais uniquement avant une nouvelle pause
              

                # Vérifier s'il s'agit de la première ou de la deuxième pause
                if self.horloge_count == 1:
                    # Compléter la pause existante avec un tiret suivi de la deuxième heure
                    self.entry_heure_pauses.insert(len(self.entry_heure_pauses.get()), f"-{heure_formatee}")
                else:
                    # Insérer la première pause
                    self.entry_heure_pauses.insert(len(self.entry_heure_pauses.get()), heure_formatee)

                # Fermer la fenêtre de sélection de l'heure
                fenetre_horloge.destroy()

                # Incrémenter le compteur
                self.horloge_count += 1

                # Si moins de deux heures ont été choisies, ouvrir une autre horloge
                if self.horloge_count < 2:
                    ouvrir_heurloge()

            fenetre_horloge = Toplevel(self)
            fenetre_horloge.title("Sélectionner l'heure")
            horloge = AnalogPicker(fenetre_horloge)
            horloge.pack(pady=10)
            btn_choisir = ctk.CTkButton(fenetre_horloge, text="Valider", command=choisir_heure)
            btn_choisir.pack(pady=10)
            theme = AnalogThemes(horloge)
            theme.setLightBlueTheme()  # Exemple : Thème "Lightblue"

        # Ajouter une virgule si le champ est déjà rempli avant d'insérer une nouvelle pause
        if self.entry_heure_pauses.get():
            self.entry_heure_pauses.insert(len(self.entry_heure_pauses.get()), ",")

        ouvrir_heurloge()





        

    def demander_fichier(self):
        self.fichier_excel = filedialog.askopenfilename(
            title="Sélectionner le fichier Excel",
            filetypes=[("Fichiers Excel", "*.xls;*.xlsx;*.xlsm;*.xlsb")]
        )
        if self.fichier_excel:
            messagebox.showinfo("Succès", f"Fichier sélectionné : {self.fichier_excel}")
        else:
            messagebox.showerror("Erreur", "Aucun fichier sélectionné.")
    def get_pauses(self):
        # Récupérer le contenu de l'entrée
        pauses_text = self.entry_heure_pauses.get()
        
        # Convertir en tableau (liste) en séparant par les virgules
        pauses_list = pauses_text.split(',')

        # Transformer chaque plage horaire en tuple (start_time, end_time)
        pauses = []
        for pause in pauses_list:
            if '-' in pause:
                start, end = pause.split('-')
                pauses.append((start.strip(), end.strip()))
            else:
                print(f"Format incorrect pour la pause: {pause}")
        
        print("Heures de pause sous forme de tuples:", pauses)
        return pauses


    def lancer_traitement(self):
        try:
            # Récupérer les valeurs de l'interface
            self.production_par_heure = float(self.entry_production.get())
            self.date_heure_debut = datetime.strptime(self.entry_date_heure.get(), "%Y-%m-%d %H:%M:%S")
            self.prefixe_fichier = self.entry_prefixe.get().strip()

            if not self.fichier_excel:
                raise ValueError("Aucun fichier Excel sélectionné.")

            # Charger et traiter le fichier Excel
            self.traiter_fichier_excel()
            self.get_pauses()

            messagebox.showinfo("Succès", "Le traitement du fichier a été effectué avec succès.")

        except ValueError as ve:
            messagebox.showerror("Erreur", f"Entrée invalide : {ve}")
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur s'est produite : {e}")
    

    

    def ajuster_pour_pauses(self, current_datetime, heures_restantes, pauses):
        """
        Ajuste le temps restant pour tenir compte des pauses.
        """
        # Calculer l'heure de fin de travail avant d'ajouter les pauses
        fin_datetime = current_datetime + timedelta(hours=heures_restantes)
        
        for pause_start, pause_end in pauses:
            # Convertir les heures de pause en datetime
            pause_debut = current_datetime.replace(hour=int(pause_start.split(":")[0]), 
                                                    minute=int(pause_start.split(":")[1]), 
                                                    second=0)
            pause_fin = current_datetime.replace(hour=int(pause_end.split(":")[0]), 
                                                minute=int(pause_end.split(":")[1]), 
                                                second=0)
            
            # Si la fin de la pause est avant le début (cela signifie que la pause traverse minuit)
            if pause_fin < pause_debut:
                # Ajouter un jour à la fin de la pause pour la gérer correctement
                pause_fin += timedelta(days=1)

            # Si la période de travail chevauche la pause
            if current_datetime < pause_fin and fin_datetime > pause_debut:
                # Ajouter la durée de la pause
                fin_datetime += pause_fin - pause_debut
        
        return fin_datetime


    def traiter_fichier_excel(self):
        """
        Charge et traite le fichier Excel, en sélectionnant automatiquement le moteur en fonction de l'extension.
        """
        # Vérifier l'extension du fichier
        _, extension = os.path.splitext(self.fichier_excel)

        # Sélectionner le moteur en fonction de l'extension
        if extension == ".xlsx" or extension == ".xlsm":
            moteur = "openpyxl"
        elif extension == ".xls":
            moteur = "xlrd"
        elif extension == ".xlsb":
            moteur = "pyxlsb"
        else:
            raise ValueError(f"Format de fichier non pris en charge : {extension}")

        # Charger le fichier avec le moteur approprié
        try:
            df = pd.read_excel(self.fichier_excel, engine=moteur)
            print(f"Fichier chargé avec succès en utilisant le moteur '{moteur}'.")
        except Exception as e:
            raise ValueError(f"Erreur lors du chargement du fichier avec le moteur '{moteur}': {e}")

        # Traitement des données
        df_principales = df[df['Déclaration\nFin Prépa ?'].isna()]
        colonnes_a_garder = ['Chargement ', 'Date Chargement', 'Nb Prépa', 'UVC à préparer', 'UVC Prépa Validées']
        df_principales = df_principales[colonnes_a_garder]

        # Supprimer les lignes où 'Chargement ' == 'Sous-Total'
        df_principales = df_principales[df_principales['Chargement ']!= 'Sous-Total']
        df_principales = df_principales[df_principales['Date Chargement'] >= self.date_heure_debut]


        # Calculer UVC restants et heures nécessaires
        df_principales['UVC Restants'] = df_principales['UVC à préparer'] - df_principales['UVC Prépa Validées']
        df_principales['Heures nécessaires'] = df_principales['UVC Restants'] / self.production_par_heure

        # Générer les heures de fin
        date_heure = self.date_heure_debut
        current_datetime = date_heure

        heures_de_fin = []
        previous_chargement = None
        pauses = self.get_pauses()
        cptT = 0

        for i, row in df_principales.iterrows():
            chargement_value = row[0]
            
            
            heures_restantes = row['Heures nécessaires']
            fin_datetime = self.ajuster_pour_pauses(current_datetime, heures_restantes, pauses)
            heures_de_fin.append(fin_datetime.strftime("%Y-%m-%d %H:%M:%S"))
            current_datetime = fin_datetime

            previous_chargement = chargement_value

        df_principales['Heures de fin'] = heures_de_fin

        # Sauvegarder le fichier traité
        fichier_temporaire = "temp_file.xlsx"
        df_principales.to_excel(fichier_temporaire, index=False)

        # Appliquer la mise en forme conditionnelle
        wb = load_workbook(fichier_temporaire)
        sheet = wb.active

        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            chargement_value = row[0].value
            date_chargement = row[1].value
            heure_de_fin = row[-1].value

            if chargement_value == "Sous-Total":
                for cell in row:
                    cell.fill = blue_fill
            elif heure_de_fin and date_chargement:
                heure_de_fin_dt = datetime.strptime(heure_de_fin, "%Y-%m-%d %H:%M:%S")
                date_chargement_dt = datetime.strptime(str(date_chargement), "%Y-%m-%d %H:%M:%S")
                
                if heure_de_fin_dt > date_chargement_dt:
                    for cell in row:
                        cell.fill = red_fill
                elif 0 <= (date_chargement_dt - heure_de_fin_dt).total_seconds() <= 1800:
                    for cell in row:
                        cell.fill = yellow_fill
        self.df_principales = df_principales

        # Ajouter le préfixe au fichier final
        prefixe = self.prefixe_fichier if self.prefixe_fichier else "Fichier_Traité"
        fichier_final = f"{prefixe}.xlsx"
        wb.save(fichier_final)
        os.remove(fichier_temporaire)
    
    def preview_fichier(self):
        """
        Ouvre une nouvelle fenêtre pour afficher un aperçu des données du fichier Excel avec une grille et des couleurs.
        Permet le défilement vertical si les données dépassent la hauteur de la fenêtre.
        """
        if not hasattr(self, 'df_principales') or self.df_principales is None:
            messagebox.showerror("Erreur", "Aucune donnée disponible pour l'aperçu.")
            return

        # Créer une nouvelle fenêtre pour l'aperçu
        preview_window = Toplevel(self)
        preview_window.title("Aperçu des données")
        preview_window.geometry("800x600")  # Ajustez cette taille selon votre besoin

        # Conteneur principal avec barre de défilement
        container = ctk.CTkFrame(preview_window)
        container.pack(fill="both", expand=True)

        canvas = ctk.CTkCanvas(container)
        canvas.pack(side="left", fill="both", expand=True)

        scrollbar = ctk.CTkScrollbar(container, orientation="vertical", command=canvas.yview)
        scrollbar.pack(side="right", fill="y")

        canvas.configure(yscrollcommand=scrollbar.set)

        # Frame intérieure pour afficher le contenu
        inner_frame = ctk.CTkFrame(canvas)
        canvas.create_window((canvas.winfo_width() // 2, canvas.winfo_height() // 2), window=inner_frame, anchor="center")


        # Créer les en-têtes de colonnes avec bordures
        for i, column in enumerate(self.df_principales.columns):
            header_frame = ctk.CTkFrame(inner_frame, border_width=1, border_color="black")  # Ajout de bordures
            header_frame.grid(row=0, column=i, sticky="nsew", padx=1, pady=1)  # Grille serrée avec espace
            label = ctk.CTkLabel(header_frame, text=column, font=("Arial", 12, "bold"), fg_color="lightblue")
            label.pack(fill="both", expand=True)

        # Définir les couleurs selon les critères
        def get_cell_color(row):
            """
            Retourne une couleur en fonction des critères spécifiques à la ligne.
            """
            chargement_value = row[0]
            heure_de_fin = row['Heures de fin']
            date_chargement = row['Date Chargement']

            # Vérifier si les valeurs sont valides avant de les utiliser
            if pd.isna(heure_de_fin) or pd.isna(date_chargement):
                return "white"  # Blanc pour les cellules non valides

            try:
                heure_de_fin_dt = datetime.strptime(heure_de_fin, "%Y-%m-%d %H:%M:%S")
                date_chargement_dt = datetime.strptime(str(date_chargement), "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return "white"  # Blanc pour les formats non valides

            if chargement_value == "Sous-Total":
                return "lightblue"
            elif heure_de_fin_dt > date_chargement_dt:
                return "red"
            elif 0 <= (date_chargement_dt - heure_de_fin_dt).total_seconds() <= 1800:
                return "yellow"
            return "white"


        # Créer les lignes de données avec bordures et couleurs
        for i, row in self.df_principales.iterrows():
            for j, value in enumerate(row):
                cell_color = get_cell_color(row)  # Appliquer la couleur à chaque cellule
                cell_frame = ctk.CTkFrame(inner_frame, border_width=1, border_color="gray", fg_color=cell_color)
                cell_frame.grid(row=i + 1, column=j, sticky="nsew", padx=1, pady=1)
                label = ctk.CTkLabel(cell_frame, text=str(value), font=("Arial", 12))
                label.pack(fill="both", expand=True)

        # Ajuster les colonnes pour qu'elles soient uniformes
        for col in range(len(self.df_principales.columns)):
            inner_frame.grid_columnconfigure(col, weight=1)

        # Centrer le tableau dans la fenêtre
        # Configuration de la grille intérieure pour occuper tout l'espace
        inner_frame.grid_rowconfigure(0, weight=1)  # Premier row a du poids
        inner_frame.grid_columnconfigure(0, weight=1)  # Premier column a du poids

        # Ajouter un padding autour du tableau pour centrer son contenu
        for row in range(1, len(self.df_principales) + 1):
            inner_frame.grid_rowconfigure(row, weight=1)
        for col in range(len(self.df_principales.columns)):
            inner_frame.grid_columnconfigure(col, weight=1)

        # Mettre à jour la taille du canevas pour ajuster au contenu
        inner_frame.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))

        # Activer le défilement avec la molette
        def on_mousewheel(event):
            canvas.yview_scroll(-1 * int(event.delta / 120), "units")

        canvas.bind_all("<MouseWheel>", on_mousewheel)

        # Bouton pour fermer la fenêtre
        button_close = ctk.CTkButton(preview_window, text="Retour", command=preview_window.destroy)
        button_close.pack(pady=10)

# Lancer l'application
if __name__ == "__main__":
    app = Application()
    app.mainloop()
